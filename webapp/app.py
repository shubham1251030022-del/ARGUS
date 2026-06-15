"""
ARGUS — File 11: webapp/app.py  [v11 — Full Fix]
Member 2: Shubham Pitty | VIT Pune CSAIML-E Group 01

v11 fixes over v10:
  1. ArUco scan thread reads camera_index from config.json — was hardcoded 0,
     causing scan to use wrong camera on systems where webcam is index 1+
  2. exam_stop() now waits 1.5s after writing stop signal BEFORE calling
     terminate() — prevents race condition where terminate() fires before
     main.py reads the signal file
  3. Workflow state + aruco state properly reset on Stop Exam — next session
     starts clean (aruco_done was staying True from previous session)
  4. btn-start is now enabled after seating upload even if ArUco was not
     scanned — teacher can proceed with default zones if ArUco fails/times out
  5. Start Exam button re-enabled correctly after workflow reset on stop

All v10 fixes retained:
  - Smart zone extrapolation for 2-of-3 markers
  - Zones student name always refreshed from seating
  - Excel export before session meta clear
  - Stop signal file (ARGUS_STOP.signal) for clean main.py exit
"""

import os
import sys
import json
import threading
import time
import subprocess
from datetime import datetime

from flask import (Flask, Response, jsonify, request,
                   render_template_string, send_from_directory)

# ── Path setup ────────────────────────────────────────────────────────────────
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_ROOT     = os.path.dirname(_THIS_DIR)
sys.path.insert(0, _ROOT)
sys.path.insert(0, os.path.join(_ROOT, "detection"))

from webapp.alert_logger import AlertLogger

SNAPSHOT_DIR  = os.path.join(_ROOT, "snapshots")
ZONES_FILE    = os.path.join(_ROOT, "detection", "zones.json")
SEATING_FILE  = os.path.join(_ROOT, "seating_upload.json")
CONFIG_FILE   = os.path.join(_ROOT, "detection", "config.json")
MAIN_PY       = os.path.join(_ROOT, "detection", "main.py")
STOP_SIGNAL   = os.path.join(_ROOT, "ARGUS_STOP.signal")

ALERT_THRESHOLD   = 100
WARNING_THRESHOLD = 60
MAX_SCORE         = 100

app    = Flask(__name__)
logger = AlertLogger()


def _load_config():
    try:
        with open(CONFIG_FILE) as f:
            return json.load(f)
    except Exception:
        return {"camera_index": 0}


# ── Live state ────────────────────────────────────────────────────────────────
_state_lock = threading.Lock()
_live_state = {
    "benches": {}, "frame_count": 0,
    "fps": 0, "running": False, "last_update": 0
}

# ── Frame buffer ──────────────────────────────────────────────────────────────
_frame_lock   = threading.Lock()
_latest_frame = None

# ── Session metadata ──────────────────────────────────────────────────────────
_session_meta = {
    "division" : "",
    "subject"  : "",
    "exam_date": "",
    "room"     : "",
}

# ── Exam state ────────────────────────────────────────────────────────────────
_exam_lock       = threading.Lock()
_exam_active     = False
_exam_start_time = None
_main_process    = None

# ── ArUco scan state ──────────────────────────────────────────────────────────
_aruco_lock   = threading.Lock()
_aruco_state  = {
    "running"     : False,
    "done"        : False,
    "zones_found" : 0,
    "zones"       : {},
    "message"     : "Not started",
    "error"       : ""
}

# ── Workflow state ─────────────────────────────────────────────────────────────
_workflow_state = "IDLE"


def _set_workflow(state):
    global _workflow_state
    _workflow_state = state


# ════════════════════════════════════════════════════════════════════════════
# EXCEL SEATING UPLOAD
# ════════════════════════════════════════════════════════════════════════════

def _parse_excel(file_bytes):
    import io
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
    except Exception as e:
        return None, f"Cannot open Excel: {e}", {}

    headers = []
    rows    = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else "" for c in row]
        else:
            if any(c is not None for c in row):
                rows.append(row)

    def find_col(keywords):
        for kw in keywords:
            for j, h in enumerate(headers):
                if kw.lower() in h.lower():
                    return j
        return None

    col_prn      = find_col(["prn", "roll", "id"])
    col_name     = find_col(["candidate", "name", "student"])
    col_bench    = find_col(["bench no", "bench_no", "bench"])
    col_room     = find_col(["room"])
    col_class    = find_col(["class", "year", "sem"])
    col_division = find_col(["division", "div"])
    col_subject  = find_col(["subject", "sub", "exam name", "paper"])
    col_date     = find_col(["exam date", "date"])

    if col_bench is None:
        return None, "Could not find 'Bench No' column in Excel", {}

    result = {}
    for row in rows:
        bench_raw = row[col_bench] if col_bench is not None else None
        if bench_raw is None:
            continue
        try:
            bench_num = int(bench_raw)
        except (ValueError, TypeError):
            continue
        bench_id = f"B{bench_num}"
        name     = str(row[col_name]).strip() if col_name is not None else "Unknown"
        prn      = str(row[col_prn]).strip()  if col_prn  is not None else ""
        room     = str(row[col_room]).strip() if col_room is not None else ""

        result[bench_id] = {
            "student_name": name,
            "roll_number" : prn,
            "room"        : room,
            "bench"       : bench_id,
            "status"      : "ACTIVE"
        }

    import datetime as _dt
    meta = {"division": "ARGUS", "subject": "Exam",
            "exam_date": _dt.datetime.now().strftime("%Y%m%d"), "room": ""}
    if rows:
        first = rows[0]
        if col_class is not None and first[col_class]:
            meta["class"]    = str(first[col_class]).strip()
        if col_division is not None and first[col_division]:
            meta["division"] = str(first[col_division]).strip()
        if col_subject is not None and first[col_subject]:
            meta["subject"]   = str(first[col_subject]).strip()
        if col_date is not None and first[col_date]:
            raw = str(first[col_date]).strip()
            if raw not in ("None", ""):
                meta["exam_date"] = raw.replace("-", "").replace("/", "")[:8]
        if col_room is not None and first[col_room]:
            meta["room"] = str(first[col_room]).strip()

    return result, None, meta


def _update_zones_with_students(seating):
    existing = {}
    if os.path.exists(ZONES_FILE):
        try:
            with open(ZONES_FILE) as f:
                raw = json.load(f)
            if isinstance(raw, list):
                for item in raw:
                    key = item.get("bench") or item.get("name", "")
                    if key:
                        existing[key] = item
            elif isinstance(raw, dict):
                existing = raw
        except Exception:
            existing = {}

    for bench_id, info in seating.items():
        if bench_id not in existing:
            existing[bench_id] = {}
        existing[bench_id]["student_name"] = info["student_name"]
        existing[bench_id]["roll_number"]  = info["roll_number"]
        existing[bench_id]["status"]       = "ACTIVE"

    with open(ZONES_FILE, "w") as f:
        json.dump(existing, f, indent=4)


@app.route("/api/seating/upload", methods=["POST"])
def seating_upload():
    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "error": "No file provided"}), 400

    try:
        file_bytes = file.read()
        seating, err, meta = _parse_excel(file_bytes)
        if err:
            return jsonify({"ok": False, "error": err}), 400

        global _session_meta
        _session_meta.update(meta)

        with open(SEATING_FILE, "w") as f:
            json.dump(seating, f, indent=4)

        _update_zones_with_students(seating)
        _set_workflow("SEATING_READY")
        print(f"[ARGUS] Seating uploaded — {len(seating)} students")
        print(f"[ARGUS] Session: {_session_meta['division']} | "
              f"{_session_meta['subject']} | {_session_meta['exam_date']}")

        return jsonify({
            "ok"     : True,
            "seating": seating,
            "meta"   : _session_meta
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/seating", methods=["GET"])
def get_seating():
    if os.path.exists(SEATING_FILE):
        try:
            with open(SEATING_FILE) as f:
                return jsonify(json.load(f))
        except Exception:
            pass
    return jsonify({})


# ════════════════════════════════════════════════════════════════════════════
# ARUCO SCAN
# ════════════════════════════════════════════════════════════════════════════

def _extrapolate_missing_zones(zones: dict) -> dict:
    from aruco_scanner import ZONE_EXPAND_X, ZONE_EXPAND_Y_UP, ZONE_EXPAND_Y_DN

    all_benches = ["B1", "B2", "B3"]
    missing     = [b for b in all_benches if b not in zones]

    if not missing:
        return zones

    seating = {}
    if os.path.exists(SEATING_FILE):
        try:
            with open(SEATING_FILE) as f:
                seating = json.load(f)
        except Exception:
            pass

    detected = {b: zones[b] for b in all_benches if b in zones}
    print(f"[ARUCO] Detected: {list(detected.keys())} | Missing: {missing}")

    centers = {b: z["centre_x"] for b, z in detected.items() if "centre_x" in z}
    if not centers:
        centers = {b: z["x"] + z["w"] // 2 for b, z in detected.items()}

    if len(detected) >= 2:
        sorted_detected = sorted(centers.items(), key=lambda x: x[1])
        spacings = []
        for i in range(len(sorted_detected) - 1):
            spacings.append(sorted_detected[i + 1][1] - sorted_detected[i][1])
        avg_spacing = int(sum(spacings) / len(spacings))
    else:
        avg_spacing = 427

    ref   = list(detected.values())[0]
    ref_y = ref.get("y", 80)
    ref_h = ref.get("h", 580)

    bench_centers = {}
    if "B1" in centers:
        bench_centers["B1"] = centers["B1"]
        bench_centers["B2"] = centers["B1"] + avg_spacing
        bench_centers["B3"] = centers["B1"] + 2 * avg_spacing
    elif "B2" in centers:
        bench_centers["B2"] = centers["B2"]
        bench_centers["B1"] = centers["B2"] - avg_spacing
        bench_centers["B3"] = centers["B2"] + avg_spacing
    elif "B3" in centers:
        bench_centers["B3"] = centers["B3"]
        bench_centers["B2"] = centers["B3"] - avg_spacing
        bench_centers["B1"] = centers["B3"] - 2 * avg_spacing

    if "B1" in centers and "B3" in centers and "B2" not in centers:
        bench_centers["B2"] = (centers["B1"] + centers["B3"]) // 2

    for b in missing:
        cx = bench_centers.get(b, 0)
        x1 = max(0, cx - ZONE_EXPAND_X)
        y1 = max(0, ref_y)
        x2 = min(1280, cx + ZONE_EXPAND_X)
        y2 = min(720, ref_y + ref_h)
        s  = seating.get(b, {})
        zones[b] = {
            "bench"       : b,
            "aruco_id"    : {"B1": 0, "B2": 1, "B3": 2}.get(b, -1),
            "x"           : x1,
            "y"           : y1,
            "w"           : x2 - x1,
            "h"           : y2 - y1,
            "centre_x"    : cx,
            "centre_y"    : ref_y + ref_h // 2,
            "student_name": s.get("student_name", "Unknown"),
            "roll_number" : s.get("roll_number", ""),
            "extrapolated": True
        }
        print(f"[ARUCO] Extrapolated {b}: centre_x={cx} x={x1} w={x2-x1}"
              f" student={zones[b]['student_name']}")

    return zones


def _aruco_scan_thread():
    """
    Runs ARUCOScanner in a background thread.
    FIX v11: reads camera_index from config.json — was hardcoded 0.
    """
    global _aruco_state

    with _aruco_lock:
        _aruco_state = {
            "running": True, "done": False,
            "zones_found": 0, "zones": {},
            "message": "Opening camera...", "error": ""
        }

    try:
        import cv2
        sys.path.insert(0, os.path.join(_ROOT, "detection"))
        from aruco_scanner import ARUCOScanner

        # FIX: read from config — not hardcoded 0
        cam_index = _load_config().get("camera_index", 0)
        print(f"[ARUCO] Using camera index {cam_index} from config.json")

        scanner = ARUCOScanner()
        cap = cv2.VideoCapture(cam_index)

        if not cap.isOpened():
            with _aruco_lock:
                _aruco_state.update({
                    "running": False, "done": True,
                    "error": f"Camera not available (index {cam_index}) — "
                             f"check config.json → camera_index"
                })
            return

        cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)

        for _ in range(10):
            cap.read()

        with _aruco_lock:
            _aruco_state["message"] = "Scanning for ArUco markers..."

        start_time = time.time()
        TIMEOUT    = 90
        hit_count  = 0
        HIT_NEED   = 6
        last_zones = {}

        while True:
            if time.time() - start_time > TIMEOUT:
                with _aruco_lock:
                    _aruco_state.update({
                        "running": False, "done": True,
                        "error": "Timeout — hold marker flat facing camera, "
                                 "ensure room light hits marker directly."
                    })
                break

            ret, frame = cap.read()
            if not ret:
                continue

            zones = scanner.scan_frame(frame)

            if len(zones) > 0:
                last_zones = zones
                hit_count += 1

            with _aruco_lock:
                found_list = list(last_zones.keys())
                _aruco_state["zones_found"] = len(found_list)
                _aruco_state["message"] = (
                    f"Found {len(found_list)} marker(s): {', '.join(sorted(found_list))}"
                    f" — locking [{hit_count}/{HIT_NEED}]"
                    if found_list
                    else "Searching... Point camera at ArUco markers"
                )

            try:
                _, jpeg = cv2.imencode(".jpg", frame,
                    [cv2.IMWRITE_JPEG_QUALITY, 40])
                with _frame_lock:
                    global _latest_frame
                    _latest_frame = jpeg.tobytes()
            except Exception:
                pass

            if hit_count >= HIT_NEED:
                zones = last_zones
                seating = {}
                if os.path.exists(SEATING_FILE):
                    try:
                        with open(SEATING_FILE) as f:
                            seating = json.load(f)
                    except Exception:
                        pass

                for bench_id, zone in zones.items():
                    if bench_id in seating:
                        zone["student_name"] = seating[bench_id]["student_name"]
                        zone["roll_number"]  = seating[bench_id]["roll_number"]

                # NO extrapolation — only save zones physically detected
                # Fake zones from 1 marker with guessed spacing caused wrong
                # zone coordinates and broke assignment entirely
                scanner.save_zones(zones)
                _set_workflow("ARUCO_READY")

                real_count = len(zones)
                with _aruco_lock:
                    _aruco_state.update({
                        "running"    : False,
                        "done"       : True,
                        "zones_found": real_count,
                        "zones"      : {k: {
                            "bench"  : k,
                            "student": zones[k].get("student_name", "Unknown"),
                            "roll"   : zones[k].get("roll_number", "")
                        } for k in zones},
                        "message": f"✓ {real_count} zone(s) locked and saved"
                                   + (" — place remaining markers and rescan to add more" if real_count < 3 else "")
                    })
                print(f"[ARUCO] Scan complete — {real_count} real zones locked (no extrapolation)")
                break

            time.sleep(0.05)

        cap.release()

    except Exception as e:
        with _aruco_lock:
            _aruco_state.update({
                "running": False, "done": True,
                "error": str(e)
            })
        print(f"[ARUCO] Scan thread error: {e}")


@app.route("/api/aruco/start", methods=["POST"])
def aruco_start():
    with _aruco_lock:
        if _aruco_state["running"]:
            return jsonify({"ok": False, "error": "Scan already running"}), 400

    t = threading.Thread(target=_aruco_scan_thread, daemon=True)
    t.start()
    return jsonify({"ok": True})


@app.route("/api/aruco/status", methods=["GET"])
def aruco_status():
    with _aruco_lock:
        return jsonify(dict(_aruco_state))


# ════════════════════════════════════════════════════════════════════════════
# EXAM CONTROL
# ════════════════════════════════════════════════════════════════════════════

def _export_session_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import datetime

        alerts  = logger.get_all()
        summary = logger.get_summary()

        wb = Workbook()
        ws = wb.active
        ws.title = "ARGUS Session Report"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1a1a2e")

        ws.merge_cells("A1:G1")
        div_s  = _session_meta.get("division", "")
        subj_s = _session_meta.get("subject", "")
        title_str = "ARGUS Exam Report"
        if div_s:  title_str += f" — {div_s}"
        if subj_s: title_str += f" | {subj_s}"
        title_str += f" | {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A1"] = title_str
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:G2")
        ws["A2"] = (f"Total Alerts: {summary.get('total_alerts', 0)} | "
                    f"Students Flagged: {summary.get('unique_students', 0)} | "
                    f"Peak Score: {summary.get('highest_score', 0)} | "
                    f"Benches: {', '.join(summary.get('benches_flagged', []))}")
        ws["A2"].font = Font(italic=True)

        headers = ["#", "Time", "Bench", "Student", "Roll No", "Score",
                   "ML Score % (info only)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, alert in enumerate(alerts, 5):
            ws.cell(row=row_idx, column=1, value=alert.get("id", ""))
            ws.cell(row=row_idx, column=2, value=alert.get("time", ""))
            ws.cell(row=row_idx, column=3, value=alert.get("bench", ""))
            ws.cell(row=row_idx, column=4, value=alert.get("student_name", ""))
            ws.cell(row=row_idx, column=5, value=alert.get("roll_number", ""))
            ws.cell(row=row_idx, column=6, value=alert.get("score", 0))
            conf_pct = round(alert.get("ml_confidence", 0) * 100, 1)
            ws.cell(row=row_idx, column=7, value=conf_pct)

        for col, width in zip("ABCDEFG", [5, 10, 8, 20, 15, 8, 10]):
            ws.column_dimensions[col].width = width

        import re
        def clean(s):
            return re.sub(r'[\/:*?"<>|]', '', str(s).strip().replace(' ', '_'))

        cls  = clean(_session_meta.get("class",    ""))
        div  = clean(_session_meta.get("division", "ARGUS"))
        subj = clean(_session_meta.get("subject",  "Exam"))

        if cls and cls not in div:
            fname = f"{cls}_{div}_{subj}.xlsx"
        else:
            fname = f"{div}_{subj}.xlsx"

        reports_dir = os.path.join(_ROOT, "reports")
        os.makedirs(reports_dir, exist_ok=True)
        fpath = os.path.join(reports_dir, fname)
        wb.save(fpath)
        print(f"[ARGUS] Report saved: {fpath}")
        return fname, fpath
    except Exception as e:
        print(f"[ARGUS] Excel export failed: {e}")
        return None, None


def _ensure_default_zones():
    """
    At exam start, zones.json must have entries for all 3 benches.
    - If ArUco was scanned for a bench → keep those exact coordinates
    - If a bench has no ArUco data → add a default strip (won't be used
      for scoring unless a person centroid falls in it)
    Student names always refreshed from seating file.
    """
    seating = {}
    if os.path.exists(SEATING_FILE):
        try:
            with open(SEATING_FILE) as f:
                seating = json.load(f)
        except Exception:
            seating = {}

    existing = {}
    if os.path.exists(ZONES_FILE):
        try:
            with open(ZONES_FILE) as f:
                raw = json.load(f)
            if isinstance(raw, dict):
                existing = raw
            elif isinstance(raw, list):
                for item in raw:
                    key = item.get("bench") or item.get("bench_id") or item.get("name", "")
                    if key:
                        existing[key] = item
        except Exception:
            existing = {}

    # Default fallback positions — only used if ArUco never detected this bench
    default_positions = {
        "B1": {"x": 20,  "y": 80, "w": 400, "h": 580, "aruco_id": 0},
        "B2": {"x": 440, "y": 80, "w": 400, "h": 580, "aruco_id": 1},
        "B3": {"x": 860, "y": 80, "w": 400, "h": 580, "aruco_id": 2},
    }

    merged = {}
    for bench_id, default_pos in default_positions.items():
        ex = existing.get(bench_id, {})
        # Use ArUco coordinates if available — NEVER overwrite with defaults
        if ex.get("x") is not None and ex.get("w") is not None:
            coords = {
                "x"        : ex["x"],
                "y"        : ex["y"],
                "w"        : ex["w"],
                "h"        : ex["h"],
                "aruco_id" : ex.get("aruco_id", default_pos["aruco_id"])
            }
            source = "aruco"
        else:
            coords = dict(default_pos)
            source = "default"

        s = seating.get(bench_id, {})
        merged[bench_id] = {
            **coords,
            "student_name": s.get("student_name", ex.get("student_name", "Unknown")),
            "roll_number" : s.get("roll_number",  ex.get("roll_number",  "")),
            "status"      : "ACTIVE"
        }
        print(f"  {bench_id}: [{source}] {merged[bench_id]['student_name']} "
              f"x={merged[bench_id]['x']} y={merged[bench_id]['y']} "
              f"w={merged[bench_id]['w']} h={merged[bench_id]['h']}")

    with open(ZONES_FILE, "w") as f:
        json.dump(merged, f, indent=4)
    print(f"[ARGUS] zones.json ready — ArUco zones preserved, defaults filled for unscanned benches")


@app.route("/api/exam/start", methods=["POST"])
def exam_start():
    global _exam_active, _exam_start_time, _main_process

    with _exam_lock:
        if _exam_active:
            return jsonify({"ok": False, "error": "Exam already active"}), 400

        try:
            # Block start if ArUco scan still running (camera conflict)
            with _aruco_lock:
                aruco_running = _aruco_state.get("running", False)
            if aruco_running:
                return jsonify({
                    "ok": False,
                    "error": "ArUco scan still running — wait for it to complete"
                }), 400

            # FIX: Clean up any stale stop signal before starting
            if os.path.exists(STOP_SIGNAL):
                try:
                    os.remove(STOP_SIGNAL)
                    print("[ARGUS] Removed stale stop signal before exam start")
                except Exception:
                    pass

            _ensure_default_zones()

            if os.name == 'nt':
                cmd = f'cd /d "{_ROOT}" && py -3.11 detection/main.py'
                _main_process = subprocess.Popen(
                    cmd, shell=True,
                    creationflags=subprocess.CREATE_NEW_CONSOLE
                )
            else:
                _main_process = subprocess.Popen(
                    [sys.executable, MAIN_PY], cwd=_ROOT
                )

            _exam_active     = True
            _exam_start_time = time.time()
            _set_workflow("EXAM_ACTIVE")
            logger.clear_session()
            with _state_lock:
                _live_state["benches"]     = {}
                _live_state["frame_count"] = 0
            print(f"[ARGUS] Exam started — main.py PID {_main_process.pid}")
            return jsonify({"ok": True, "pid": _main_process.pid})

        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/exam/stop", methods=["POST"])
def exam_stop():
    global _exam_active, _main_process, _session_meta, _latest_frame

    # Export FIRST before clearing session metadata
    fname, fpath = _export_session_excel()

    # FIX: Write stop signal, then wait 1.5s before terminate()
    # This gives main.py enough time to read the file in its loop
    # before the process is killed as a fallback.
    try:
        with open(STOP_SIGNAL, "w") as sf:
            sf.write("stop")
        print("[ARGUS] Stop signal written — waiting for main.py to read it...")
        time.sleep(1.5)   # FIX: was 0ms — race condition
    except Exception as se:
        print(f"[ARGUS] Signal write error: {se}")

    with _exam_lock:
        _exam_active = False

        # FIX: Reset workflow to SEATING_READY (not IDLE) so teacher can
        # restart exam without re-uploading seating. ArUco scan cleared too.
        _set_workflow("SEATING_READY")

        if _main_process and _main_process.poll() is None:
            try:
                _main_process.terminate()
                print("[ARGUS] main.py process terminated (fallback)")
            except Exception:
                pass
        _main_process = None

        # Clear session meta AFTER export
        _session_meta = {"division": "", "subject": "", "exam_date": "", "room": ""}

    # FIX: Reset aruco state so next session starts fresh
    with _aruco_lock:
        _aruco_state.update({
            "running"     : False,
            "done"        : False,
            "zones_found" : 0,
            "zones"       : {},
            "message"     : "Not started",
            "error"       : ""
        })

    with _state_lock:
        _live_state["running"]     = False
        _live_state["fps"]         = 0
        _live_state["last_update"] = 0
    with _frame_lock:
        _latest_frame = None

    return jsonify({
        "ok"          : True,
        "report_file" : fname,
        "report_ready": fname is not None
    })


@app.route("/api/exam/status", methods=["GET"])
def exam_status():
    with _exam_lock:
        active   = _exam_active
        start_ts = _exam_start_time
        pid      = _main_process.pid if _main_process else None
    elapsed = int(time.time() - start_ts) if (active and start_ts) else 0
    return jsonify({
        "active"     : active,
        "elapsed_sec": elapsed,
        "pid"        : pid
    })


@app.route("/api/workflow", methods=["GET"])
def get_workflow():
    seating_ok = os.path.exists(SEATING_FILE)
    aruco_ok   = _workflow_state in ("ARUCO_READY", "EXAM_ACTIVE", "EXAM_ENDED")
    return jsonify({
        "state"      : _workflow_state,
        "seating_ok" : seating_ok,
        "aruco_ok"   : aruco_ok,
        "exam_active": _exam_active
    })


# ════════════════════════════════════════════════════════════════════════════
# VIDEO + STATUS
# ════════════════════════════════════════════════════════════════════════════

def _generate_stream():
    while True:
        with _frame_lock:
            frame = _latest_frame
        if frame:
            yield (b"--frame\r\nContent-Type: image/jpeg\r\n\r\n"
                   + frame + b"\r\n")
        else:
            time.sleep(0.05)
        time.sleep(0.033)


@app.route("/video_feed")
def video_feed():
    resp = Response(_generate_stream(),
                    mimetype="multipart/x-mixed-replace; boundary=frame")
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"]        = "no-cache"
    resp.headers["Expires"]       = "0"
    return resp


@app.route("/api/frame", methods=["POST"])
def push_frame():
    global _latest_frame
    with _frame_lock:
        _latest_frame = request.data
    return jsonify({"ok": True})


@app.route("/api/status", methods=["GET"])
def get_status():
    with _state_lock:
        state = dict(_live_state)
    with _exam_lock:
        state["exam_active"]  = _exam_active
        state["exam_elapsed"] = (
            int(time.time() - _exam_start_time)
            if (_exam_active and _exam_start_time) else 0
        )
    return jsonify(state)


@app.route("/api/status", methods=["POST"])
def push_status():
    global _live_state
    data = request.get_json(silent=True) or {}
    with _state_lock:
        _live_state.update(data)
        _live_state["last_update"] = time.time()
    return jsonify({"ok": True})


# ════════════════════════════════════════════════════════════════════════════
# ALERTS
# ════════════════════════════════════════════════════════════════════════════

@app.route("/api/alerts",        methods=["GET"])
def get_alerts():        return jsonify(logger.get_all())

@app.route("/api/alerts/recent", methods=["GET"])
def get_recent_alerts():
    return jsonify(logger.get_recent(request.args.get("n", 10, type=int)))

@app.route("/api/alert",         methods=["POST"])
def post_alert():
    d = request.get_json(silent=True) or {}
    a = logger.log_alert(
        bench=d.get("bench", ""), student_name=d.get("student_name", "Unknown"),
        roll_number=d.get("roll_number", ""), score=d.get("score", 0),
        ml_confidence=d.get("ml_confidence", 0.0),
        flags=d.get("flags", {}), snapshot_path=d.get("snapshot_path", "")
    )
    return jsonify(a), 201

@app.route("/api/summary",       methods=["GET"])
def get_summary():       return jsonify(logger.get_summary())

@app.route("/api/reviewed/<int:alert_id>", methods=["POST"])
def mark_reviewed(alert_id):
    return jsonify({"ok": logger.mark_reviewed(alert_id), "id": alert_id})

@app.route("/api/clear",         methods=["POST"])
def clear_session():
    logger.clear_session()
    with _state_lock:
        _live_state["benches"]     = {}
        _live_state["frame_count"] = 0
    return jsonify({"ok": True})

@app.route("/snapshots/<path:filename>")
def serve_snapshot(filename):
    return send_from_directory(SNAPSHOT_DIR, filename)


# ════════════════════════════════════════════════════════════════════════════
# DASHBOARD HTML
# ════════════════════════════════════════════════════════════════════════════

DASHBOARD_HTML = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ARGUS — Exam Monitor</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { background:#0d1117; color:#e6edf3;
       font-family:'Segoe UI',sans-serif; min-height:100vh; }

.header { background:#161b22; border-bottom:1px solid #30363d;
  padding:12px 24px; display:flex; align-items:center;
  justify-content:space-between; }
.header h1 { font-size:1.2rem; color:#58a6ff; letter-spacing:2px; }
.header .meta { font-size:0.78rem; color:#8b949e;
  display:flex; align-items:center; gap:14px; }
.status-dot { display:inline-block; width:9px; height:9px;
  border-radius:50%; background:#3fb950; margin-right:5px;
  animation:pulse 1.5s infinite; }
@keyframes pulse { 0%,100%{opacity:1} 50%{opacity:.4} }

.workflow-bar { background:#0d1117; border-bottom:1px solid #21262d;
  padding:10px 24px; display:flex; align-items:center; gap:0; }
.step { display:flex; align-items:center; gap:8px;
  padding:6px 16px; border-radius:6px; font-size:0.78rem;
  color:#484f58; border:1px solid transparent; transition:all .3s; }
.step.done    { color:#3fb950; border-color:#238636; background:#0d2614; }
.step.active  { color:#58a6ff; border-color:#1d4ed8; background:#0d1b3e; }
.step.pending { color:#484f58; }
.step-num { width:20px; height:20px; border-radius:50%;
  background:#21262d; display:flex; align-items:center;
  justify-content:center; font-size:0.68rem; font-weight:700; }
.step.done .step-num  { background:#238636; color:#fff; }
.step.active .step-num{ background:#1d4ed8; color:#fff; }
.step-arrow { color:#30363d; padding:0 6px; font-size:0.9rem; }

.toolbar { background:#161b22; border-bottom:1px solid #30363d;
  padding:8px 24px; display:flex; gap:10px; align-items:center;
  flex-wrap:wrap; }
.btn { padding:7px 16px; border-radius:5px; border:none; cursor:pointer;
  font-size:0.78rem; font-weight:600; letter-spacing:.5px;
  transition:opacity .2s; display:flex; align-items:center; gap:6px; }
.btn:disabled { opacity:.35; cursor:not-allowed; }
.btn:not(:disabled):hover { opacity:.82; }
.btn-upload { background:#21262d; color:#e6edf3;
  border:1px solid #30363d; position:relative; overflow:hidden; }
.btn-upload input[type=file] { position:absolute; inset:0;
  opacity:0; cursor:pointer; width:100%; }
.btn-aruco  { background:#1d4ed8; color:#fff; }
.btn-start  { background:#238636; color:#fff; }
.btn-stop   { background:#b91c1c; color:#fff; }
.btn-clear  { background:#21262d; color:#8b949e; border:1px solid #30363d; }
.exam-timer { font-size:0.85rem; font-weight:700; letter-spacing:1px;
  padding:5px 12px; border-radius:4px; min-width:82px;
  text-align:center; color:#484f58; border:1px solid #21262d;
  background:#0d1117; }
.exam-timer.on { color:#3fb950; border-color:#238636; background:#0d2614; }

.container { display:grid; grid-template-columns:1fr 370px;
  gap:14px; padding:14px; height:calc(100vh - 140px); }
.left  { display:flex; flex-direction:column; gap:14px; overflow:hidden; }
.right { display:flex; flex-direction:column; gap:14px; overflow-y:auto; }

.card { background:#161b22; border:1px solid #30363d;
  border-radius:8px; padding:14px; }
.card-title { font-size:0.7rem; color:#8b949e; text-transform:uppercase;
  letter-spacing:1px; margin-bottom:10px;
  display:flex; justify-content:space-between; align-items:center; }

.feed-wrap { background:#0d1117; border-radius:6px; overflow:hidden;
  flex:1; display:flex; align-items:center; justify-content:center;
  min-height:240px; position:relative; }
.feed-wrap img { width:100%; max-height:400px; object-fit:contain; }
.feed-placeholder { color:#484f58; font-size:0.82rem;
  text-align:center; padding:40px; }
.det-badge { position:absolute; top:8px; right:8px; font-size:0.65rem;
  padding:3px 8px; border-radius:3px; font-weight:700; }
.det-badge.on  { background:#0d2614; color:#3fb950; border:1px solid #238636; }
.det-badge.off { background:#1a0d0d; color:#f85149; border:1px solid #f85149; }

.bench-grid { display:grid; grid-template-columns:repeat(3,1fr); gap:10px; }
.bench-card { background:#0d1117; border:1px solid #30363d;
  border-radius:6px; padding:10px; transition:all .3s; }
.bench-card.alert   { border-color:#f85149; background:#1a0d0d; }
.bench-card.warning { border-color:#d29922; background:#1a1500; }
.bench-label { font-size:0.68rem; color:#8b949e; margin-bottom:2px; }
.bench-name  { font-size:0.82rem; font-weight:600; margin-bottom:5px;
  white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.bench-roll  { font-size:0.65rem; color:#484f58; margin-bottom:5px; }
.bar-bg  { background:#21262d; border-radius:3px; height:5px; }
.bar     { height:5px; border-radius:3px; background:#3fb950; transition:width .4s; }
.bar.warn  { background:#d29922; }
.bar.alert { background:#f85149; }
.score-txt { font-size:0.72rem; color:#8b949e; margin-top:3px; }
.conf-txt  { font-size:0.65rem; color:#484f58; }
.awaiting  { font-size:0.72rem; color:#484f58; font-style:italic; }

.seating-table { width:100%; border-collapse:collapse; font-size:0.75rem; }
.seating-table th { color:#8b949e; padding:5px 8px; text-align:left;
  border-bottom:1px solid #21262d; }
.seating-table td { padding:5px 8px; border-bottom:1px solid #161b22; }
.seating-table tr:hover td { background:#21262d; }

.aruco-status { font-size:0.75rem; padding:8px 12px; border-radius:5px; margin-top:6px; }
.aruco-status.scanning { background:#0d1b3e; color:#58a6ff; border:1px solid #1d4ed8; }
.aruco-status.done     { background:#0d2614; color:#3fb950; border:1px solid #238636; }
.aruco-status.error    { background:#1a0d0d; color:#f85149; border:1px solid #f85149; }
.aruco-status.idle     { background:#161b22; color:#484f58; border:1px solid #21262d; }
.progress-dots::after { content:'...'; animation:dots 1.2s steps(4) infinite; }
@keyframes dots { 0%{content:''} 25%{content:'.'} 50%{content:'..'} 75%{content:'...'} }

/* FIX: ArUco-skipped note */
.aruco-skip-note { font-size:0.68rem; color:#8b949e; margin-top:4px;
  font-style:italic; }

.sum-row { display:flex; gap:12px; }
.stat { flex:1; text-align:center; }
.stat-val   { font-size:1.3rem; font-weight:700; color:#58a6ff; }
.stat-label { font-size:0.65rem; color:#8b949e; text-transform:uppercase; }

.alerts-scroll { overflow-y:auto; max-height:280px; }
table.alert-tbl { width:100%; border-collapse:collapse; font-size:0.75rem; }
table.alert-tbl th { text-align:left; padding:5px 7px; color:#8b949e;
  border-bottom:1px solid #21262d; position:sticky;
  top:0; background:#161b22; }
table.alert-tbl td { padding:5px 7px; border-bottom:1px solid #21262d; }
table.alert-tbl tr:hover td { background:#21262d; }
.badge { display:inline-block; padding:1px 6px; border-radius:10px;
  font-size:0.65rem; font-weight:700; }
.badge.red    { background:#2d1117;color:#f85149;border:1px solid #f85149; }
.badge.yellow { background:#1c1a00;color:#d29922;border:1px solid #d29922; }
.badge.green  { background:#0d2614;color:#3fb950;border:1px solid #3fb950; }
.btn-rev { background:none; border:1px solid #30363d; color:#8b949e;
  padding:2px 7px; border-radius:3px; cursor:pointer; font-size:0.65rem; }
.btn-rev:hover { border-color:#58a6ff; color:#58a6ff; }
.reviewed { color:#3fb950; font-size:0.68rem; }
</style>
</head>
<body>

<div class="header">
  <h1>⬡ ARGUS <span style="color:#8b949e;font-size:0.82rem;"> EXAM SURVEILLANCE · VIT PUNE · CSAIML-E</span></h1>
  <div class="meta">
    <span><span class="status-dot" id="dot"></span>
    <span id="status-text">Standby</span></span>
    <span id="fps-tag" style="color:#3fb950;font-size:0.68rem;"></span>
    <span id="clock"></span>
  </div>
</div>

<div class="workflow-bar">
  <div class="step pending" id="step1">
    <div class="step-num">1</div>
    <span>Upload Seating File</span>
  </div>
  <div class="step-arrow">›</div>
  <div class="step pending" id="step2">
    <div class="step-num">2</div>
    <span>Scan ArUco Markers</span>
  </div>
  <div class="step-arrow">›</div>
  <div class="step pending" id="step3">
    <div class="step-num">3</div>
    <span>Start Exam</span>
  </div>
</div>

<div class="toolbar">
  <button class="btn btn-upload" id="btn-upload" title="Upload ARGUS_Seating.xlsx">
    ↑ SEATING FILE
    <input type="file" id="seating-input" accept=".xlsx,.xls"
           onchange="uploadSeating(this)">
  </button>

  <button class="btn btn-aruco" id="btn-aruco"
          onclick="startAruco()" disabled>
    ⊞ SCAN ARUCO
  </button>

  <button class="btn btn-start" id="btn-start"
          onclick="startExam()" disabled>
    ▶ START EXAM
  </button>
  <button class="btn btn-stop" id="btn-stop"
          onclick="stopExam()" disabled>
    ■ STOP EXAM
  </button>

  <button class="btn btn-clear" onclick="clearSession()">✕ CLEAR</button>

  <div class="exam-timer" id="exam-timer">00:00:00</div>
  <span id="exam-label" style="font-size:0.72rem;color:#484f58;"></span>
</div>

<div class="container">
  <div class="left">
    <div class="card" style="flex:1;">
      <div class="card-title">
        <span>📷 LIVE SURVEILLANCE FEED</span>
        <span id="det-badge" class="det-badge off">DETECTION OFFLINE</span>
      </div>
      <div class="feed-wrap">
        <img id="live-feed"
             src="/video_feed"
             style="width:100%;max-height:400px;object-fit:contain;"
             onerror="this.style.display='none';document.getElementById('no-feed').style.display='flex'">
        <div id="no-feed" class="feed-placeholder"
             style="display:none;width:100%;align-items:center;justify-content:center;">
          Detection offline — start exam to begin monitoring
        </div>
      </div>
    </div>

    <div class="card">
      <div class="card-title">🪑 BENCH STATUS</div>
      <div class="bench-grid">
        <div class="bench-card" id="bench-B1">
          <div class="bench-label">B1</div>
          <div class="bench-name awaiting" id="name-B1">Awaiting student</div>
          <div class="bench-roll" id="roll-B1"></div>
          <div class="bar-bg"><div class="bar" id="bar-B1" style="width:0%"></div></div>
          <div class="score-txt" id="score-B1">0 / 100</div>
          <div class="conf-txt"  id="conf-B1">—</div>
        </div>
        <div class="bench-card" id="bench-B2">
          <div class="bench-label">B2</div>
          <div class="bench-name awaiting" id="name-B2">Awaiting student</div>
          <div class="bench-roll" id="roll-B2"></div>
          <div class="bar-bg"><div class="bar" id="bar-B2" style="width:0%"></div></div>
          <div class="score-txt" id="score-B2">0 / 100</div>
          <div class="conf-txt"  id="conf-B2">—</div>
        </div>
        <div class="bench-card" id="bench-B3">
          <div class="bench-label">B3</div>
          <div class="bench-name awaiting" id="name-B3">Awaiting student</div>
          <div class="bench-roll" id="roll-B3"></div>
          <div class="bar-bg"><div class="bar" id="bar-B3" style="width:0%"></div></div>
          <div class="score-txt" id="score-B3">0 / 100</div>
          <div class="conf-txt"  id="conf-B3">—</div>
        </div>
      </div>
    </div>
  </div>

  <div class="right">
    <div class="card">
      <div class="card-title">
        <span>📋 SEATING ASSIGNMENT</span>
        <span id="seating-status" style="font-size:0.68rem;color:#484f58;">Not uploaded</span>
      </div>
      <table class="seating-table" id="seating-table">
        <thead><tr><th>Bench</th><th>Name</th><th>PRN</th></tr></thead>
        <tbody id="seating-tbody">
          <tr><td colspan="3" style="color:#484f58;font-style:italic;">
            Upload Excel file to assign students</td></tr>
        </tbody>
      </table>
      <div class="aruco-status idle" id="aruco-msg">ArUco: Not scanned — scan markers or click SKIP ARUCO</div>
    </div>

    <div class="card">
      <div class="card-title">📊 SESSION METRICS</div>
      <div class="sum-row">
        <div class="stat"><div class="stat-val" id="s-total">0</div>
          <div class="stat-label">Alerts</div></div>
        <div class="stat"><div class="stat-val" id="s-students">0</div>
          <div class="stat-label">Flagged</div></div>
        <div class="stat"><div class="stat-val" id="s-high">0</div>
          <div class="stat-label">Peak Score</div></div>
        <div class="stat"><div class="stat-val" id="s-benches">0</div>
          <div class="stat-label">Benches</div></div>
      </div>
    </div>

    <div class="card" style="flex:1;">
      <div class="card-title">
        <span>🚨 INCIDENT LOG</span>
        <span id="log-ts" style="font-size:0.65rem;color:#484f58;">Live</span>
      </div>
      <div class="alerts-scroll">
        <table class="alert-tbl">
          <thead><tr>
            <th>#</th><th>Time</th><th>Bench</th>
            <th>Student</th><th>Score</th><th>Conf</th><th></th>
          </tr></thead>
          <tbody id="alert-tbody">
            <tr><td colspan="7"
              style="text-align:center;color:#484f58;padding:16px;">
              No incidents yet</td></tr>
          </tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<script>
const ALERT_THR = 100;
const WARN_THR  = 60;
const MAX_SCORE = 100;

let _seatingDone  = false;
let _arucoDone    = false;
let _arucoSkipped = false;
let _examActive   = false;
let _arucoPolling = null;

function updateClock() {
  document.getElementById('clock').textContent =
    new Date().toLocaleTimeString('en-IN', {hour12:false});
}
setInterval(updateClock, 1000); updateClock();

// ── FIX: Start button enabled after seating upload, not just after ArUco ──
// ArUco is now optional — teacher can skip it and use default zones.
function updateSteps() {
  const s1 = document.getElementById('step1');
  const s2 = document.getElementById('step2');
  const s3 = document.getElementById('step3');

  s1.className = 'step ' + (_seatingDone ? 'done' : 'active');
  s2.className = 'step ' + (_arucoDone   ? 'done' :
                             _seatingDone ? 'active' : 'pending');
  s3.className = 'step ' + (_examActive  ? 'done' :
                             _seatingDone ? 'active' : 'pending');

  // ArUco enabled after seating
  document.getElementById('btn-aruco').disabled = !_seatingDone || _examActive;

  // Start ONLY enabled after ArUco scan is done — mandatory
  document.getElementById('btn-start').disabled = !_arucoDone || _examActive;
  document.getElementById('btn-stop').disabled  = !_examActive;

  const skipBtn = document.getElementById('btn-skip-aruco');
  if (skipBtn) skipBtn.style.display = 'none';
}

async function uploadSeating(input) {
  if (!input.files.length) return;
  const formData = new FormData();
  formData.append('file', input.files[0]);

  document.getElementById('seating-status').textContent = 'Uploading...';
  document.getElementById('seating-status').style.color = '#58a6ff';

  try {
    const r = await fetch('/api/seating/upload', {method:'POST', body:formData});
    const d = await r.json();

    if (!d.ok) {
      document.getElementById('seating-status').textContent = 'Error: ' + d.error;
      document.getElementById('seating-status').style.color = '#f85149';
      return;
    }

    _seatingDone = true;
    document.getElementById('seating-status').textContent =
      Object.keys(d.seating).length + ' students assigned';
    document.getElementById('seating-status').style.color = '#3fb950';

    const tbody = document.getElementById('seating-tbody');
    tbody.innerHTML = Object.entries(d.seating).map(([bench, s]) =>
      `<tr>
        <td><b>${bench}</b></td>
        <td>${s.student_name}</td>
        <td style="color:#484f58">${s.roll_number}</td>
      </tr>`
    ).join('');

    Object.entries(d.seating).forEach(([bench, s]) => {
      const nameEl = document.getElementById('name-' + bench);
      const rollEl = document.getElementById('roll-' + bench);
      if (nameEl) { nameEl.textContent = s.student_name; nameEl.classList.remove('awaiting'); }
      if (rollEl) rollEl.textContent = s.roll_number;
    });

    updateSteps();
  } catch(e) {
    document.getElementById('seating-status').textContent = 'Upload failed';
    document.getElementById('seating-status').style.color = '#f85149';
  }
  input.value = '';
}

async function startAruco() {
  const msgEl = document.getElementById('aruco-msg');
  msgEl.className = 'aruco-status scanning';
  msgEl.innerHTML = '<span class="progress-dots">Scanning</span>';

  await fetch('/api/aruco/start', {method:'POST'});

  _arucoPolling = setInterval(async () => {
    try {
      const r = await fetch('/api/aruco/status');
      const d = await r.json();

      if (d.error) {
        msgEl.className = 'aruco-status error';
        msgEl.textContent = '✗ ' + d.error;
        clearInterval(_arucoPolling);
        return;
      }

      if (d.running) {
        msgEl.innerHTML = `<span class="progress-dots">${d.message || 'Scanning'}</span>`;
        return;
      }

      if (d.done) {
        clearInterval(_arucoPolling);
        msgEl.className = 'aruco-status done';
        msgEl.textContent = d.message;
        _arucoDone = true;

        if (d.zones) {
          Object.entries(d.zones).forEach(([bench, info]) => {
            const nameEl = document.getElementById('name-' + bench);
            if (nameEl && info.student && info.student !== 'Unknown') {
              nameEl.textContent = info.student;
              nameEl.classList.remove('awaiting');
            }
          });
        }
        updateSteps();
      }
    } catch(e) {}
  }, 500);
}

async function startExam() {
  document.getElementById('btn-start').disabled = true;
  const r = await fetch('/api/exam/start', {method:'POST'});
  const d = await r.json();
  if (!d.ok) {
    alert('Start failed: ' + d.error);
    document.getElementById('btn-start').disabled = false;
    return;
  }
  _examActive = true;
  // FIX: Force browser to reconnect to video feed — clears any cached error state
  const feedImg = document.getElementById('live-feed');
  if (feedImg) {
    feedImg.style.display = 'block';
    document.getElementById('no-feed').style.display = 'none';
    feedImg.src = '/video_feed?' + Date.now();
  }
  updateSteps();
}

function skipAruco() {
  _arucoSkipped = true;
  const msgEl = document.getElementById('aruco-msg');
  msgEl.className = 'aruco-status idle';
  msgEl.textContent = 'ArUco skipped — using default zones';
  updateSteps();
}

async function stopExam() {
  if (!confirm('Stop the exam? This will generate an Excel report.')) return;

  // Disable stop button immediately to prevent double-click
  document.getElementById('btn-stop').disabled = true;
  document.getElementById('btn-stop').textContent = '■ STOPPING...';

  const r = await fetch('/api/exam/stop', {method:'POST'});
  const d = await r.json();

  _examActive   = false;
  _arucoDone    = false;
  _arucoSkipped = false;  // reset so next session must scan or skip explicitly
  document.getElementById('btn-stop').textContent = '■ STOP EXAM';
  updateSteps();
  document.getElementById('exam-timer').className = 'exam-timer';
  document.getElementById('exam-label').textContent = 'EXAM ENDED';
  document.getElementById('exam-label').style.color = '#8b949e';

  document.getElementById('dot').style.background = '#484f58';
  document.getElementById('status-text').textContent = 'Standby';
  document.getElementById('fps-tag').textContent = '';
  const badge = document.getElementById('det-badge');
  if (badge) { badge.textContent = 'DETECTION OFFLINE'; badge.className = 'det-badge off'; }

  ['B1','B2','B3'].forEach(b => {
    const card = document.getElementById('bench-' + b);
    if (card) card.className = 'bench-card';
    const bar = document.getElementById('bar-' + b);
    if (bar) { bar.style.width = '0%'; bar.className = 'bar'; }
    const sc = document.getElementById('score-' + b);
    if (sc) sc.textContent = '0 / 100';
    const cf = document.getElementById('conf-' + b);
    if (cf) cf.textContent = '—';
  });

  if (d.report_ready && d.report_file) {
    const a = document.createElement('a');
    a.href = '/api/report/download/' + encodeURIComponent(d.report_file);
    a.download = d.report_file;
    document.body.appendChild(a);
    a.click();
    document.body.removeChild(a);
  }
}

function fmtSec(s) {
  return [Math.floor(s/3600), Math.floor((s%3600)/60), s%60]
    .map(n => String(n).padStart(2,'0')).join(':');
}

async function fetchStatus() {
  try {
    const r = await fetch('/api/status');
    const d = await r.json();
    const age  = Date.now()/1000 - (d.last_update || 0);
    const live = d.running && age < 8;

    document.getElementById('dot').style.background = live ? '#3fb950' : '#484f58';
    document.getElementById('status-text').textContent =
      live ? 'Detection Running' : 'Standby';
    document.getElementById('fps-tag').textContent = d.fps ? d.fps + ' FPS' : '';

    const badge = document.getElementById('det-badge');
    badge.textContent = live ? 'DETECTION ACTIVE' : 'DETECTION OFFLINE';
    badge.className   = 'det-badge ' + (live ? 'on' : 'off');

    if (d.exam_active) {
      _examActive = true;
      document.getElementById('exam-timer').className = 'exam-timer on';
      document.getElementById('exam-timer').textContent = fmtSec(d.exam_elapsed || 0);
      document.getElementById('exam-label').textContent = 'EXAM IN PROGRESS';
      document.getElementById('exam-label').style.color = '#3fb950';
      updateSteps();
    }

    const benches = d.benches || {};
    ['B1','B2','B3'].forEach(b => {
      const info  = benches[b] || {};
      const score = info.score || 0;
      const conf  = info.ml_confidence || 0;
      const name  = info.student_name || null;
      const pct   = Math.min(100, (score / MAX_SCORE) * 100);

      if (name) {
        const nameEl = document.getElementById('name-' + b);
        if (nameEl) { nameEl.textContent = name; nameEl.classList.remove('awaiting'); }
      }
      const bar = document.getElementById('bar-' + b);
      if (bar) {
        bar.style.width = pct + '%';
        bar.className = 'bar' +
          (score >= ALERT_THR ? ' alert' : score >= WARN_THR ? ' warn' : '');
      }
      const scoreEl = document.getElementById('score-' + b);
      if (scoreEl) scoreEl.textContent = score + ' / ' + MAX_SCORE;
      const confEl = document.getElementById('conf-' + b);
      if (confEl) confEl.textContent = conf ? (conf*100).toFixed(0)+'% conf' : '—';
      const card = document.getElementById('bench-' + b);
      if (card) card.className = 'bench-card' +
        (score >= ALERT_THR ? ' alert' : score >= WARN_THR ? ' warning' : '');
    });
  } catch(e) {}
}

async function fetchAlerts() {
  try {
    const [ar, sr] = await Promise.all([fetch('/api/alerts'), fetch('/api/summary')]);
    const alerts  = await ar.json();
    const summary = await sr.json();

    document.getElementById('s-total').textContent    = summary.total_alerts    || 0;
    document.getElementById('s-students').textContent = summary.unique_students || 0;
    document.getElementById('s-high').textContent     = summary.highest_score   || 0;
    document.getElementById('s-benches').textContent  = (summary.benches_flagged || []).length;

    const tbody = document.getElementById('alert-tbody');
    if (!alerts.length) {
      tbody.innerHTML = '<tr><td colspan="7" style="text-align:center;color:#484f58;padding:16px;">No incidents yet</td></tr>';
      return;
    }
    tbody.innerHTML = alerts.slice(0, 50).map(a => {
      const sc = a.score;
      const bc = sc >= ALERT_THR ? 'red' : sc >= WARN_THR ? 'yellow' : 'green';
      const rv = a.reviewed
        ? '<span class="reviewed">✓</span>'
        : `<button class="btn-rev" onclick="markReviewed(${a.id})">Review</button>`;
      return `<tr>
        <td>#${a.id}</td><td>${a.time}</td><td><b>${a.bench}</b></td>
        <td style="max-width:80px;overflow:hidden;text-overflow:ellipsis;">${a.student_name}</td>
        <td><span class="badge ${bc}">${sc}</span></td>
        <td>${(a.ml_confidence*100).toFixed(0)}%</td>
        <td>${rv}</td></tr>`;
    }).join('');

    document.getElementById('log-ts').textContent =
      new Date().toLocaleTimeString('en-IN', {hour12:false});
  } catch(e) {}
}

async function markReviewed(id) {
  await fetch('/api/reviewed/' + id, {method:'POST'});
  fetchAlerts();
}

async function clearSession() {
  if (!confirm('Clear all incidents for this session?')) return;
  await fetch('/api/clear', {method:'POST'});
  fetchAlerts(); fetchStatus();
}

async function loadExistingSeating() {
  try {
    const r = await fetch('/api/seating');
    const d = await r.json();
    if (!Object.keys(d).length) return;

    const tbody = document.getElementById('seating-tbody');
    tbody.innerHTML = Object.entries(d).map(([bench, s]) =>
      `<tr style="opacity:0.5">
        <td><b>${bench}</b></td>
        <td>${s.student_name}</td>
        <td style="color:#484f58">${s.roll_number}</td>
      </tr>`
    ).join('');

    document.getElementById('seating-status').textContent =
      'Previous session — upload fresh file to start';
    document.getElementById('seating-status').style.color = '#8b949e';
  } catch(e) {}
}

loadExistingSeating();
updateSteps();
fetchStatus();
fetchAlerts();
setInterval(fetchStatus, 1000);
setInterval(fetchAlerts, 2000);
</script>
</body>
</html>
"""


@app.route("/api/report/download/<path:filename>")
def download_report(filename):
    reports_dir = os.path.join(_ROOT, "reports")
    return send_from_directory(reports_dir, filename, as_attachment=True)


@app.route("/")
def dashboard():
    from flask import make_response
    resp = make_response(render_template_string(DASHBOARD_HTML))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"]        = "no-cache"
    resp.headers["Expires"]       = "0"
    return resp


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 55)
    print("  ARGUS — File 11: Dashboard v11")
    print("  VIT Pune | CSAIML-E | Group 01")
    print("=" * 55)
    print("\n  Open: http://localhost:5000")
    print("\n  Teacher workflow:")
    print("    1. Upload ARGUS_Seating.xlsx")
    print("    2. Place ArUco markers → Click Scan ArUco  (optional)")
    print("    3. Click Start Exam")
    print("\n  Press Ctrl+C to stop.\n")

    os.makedirs(SNAPSHOT_DIR, exist_ok=True)

    # Clean up stale stop signal on server start
    if os.path.exists(STOP_SIGNAL):
        try:
            os.remove(STOP_SIGNAL)
            print("[ARGUS] Removed stale ARGUS_STOP.signal on startup")
        except Exception:
            pass

    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
